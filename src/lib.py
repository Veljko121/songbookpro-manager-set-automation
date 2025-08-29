import requests
import openpyxl as xl
from song import Song
from set_item import SetItem
from songbookpro_manager_service import SongbookProManagerService

keys = {
    "A"  :  0,
    "B"  :  1,
    "H"  :  2,
    "C"  :  3,
    "Db" :  4, "C#" :  4,
    "D"  :  5,
    "Eb" :  6, "D#" :  6,
    "E"  :  7,
    "F"  :  8,
    "F#" :  9, "Gb" :  9,
    "G"  : 10,
    "Ab" : 11, "G#" : 11, 
    "F#m": 12, "Gbm": 12,
    "Gm" : 13,
    "G#m": 14, "Abm": 14,
    "Am" : 15,
    "Bm" : 16,
    "Hm" : 17,
    "Cm" : 18,
    "C#m": 19, "Dbm": 19,
    "Dm" : 20,
    "D#m": 21, "Ebm": 21,
    "Em" : 22,
    "Fm" : 23,

    "None" : -1,
}
    
# TODO: Refactor into LocalSheetReader
def get_sheet_names(spreadsheet: str):
    """Fetch all sheet names from the given spreadsheet."""
    return xl.load_workbook(spreadsheet, read_only=True).sheetnames

# TODO: Refactor into LocalSheetReader
def load_songs_from_sheets(spreadsheet: str, sheet_name: str):
    """
    Load songs from the given spreadsheet and sheet name.
    ## Exceptions
    - FileNotFoundError
    - KeyError
    - ValueError
    """
    doc = xl.load_workbook(spreadsheet)
    sheet = doc[sheet_name]
    
    songs = []
    row = 1
    while True:
        name_cell = sheet.cell(row, 2).value
        if name_cell is None:
            break
        
        name = str(name_cell).replace("‘", "'").replace("’", "'")
        key_cell = sheet.cell(row, 3).value
        if key_cell is None:
            raise ValueError(f"Missing key for song '{name}' in row {row}")
        
        key_str = str(key_cell).strip()
        if key_str not in keys:
            raise KeyError(f"Unknown key '{key_str}' for song '{name}' in row {row}")
        
        key_value = keys[key_str]
        songs.append((name.strip(), key_value))
        row += 1

    return songs

def match_songs(all_songs, sheet_songs):
    matched_songs = []
    for order, sheet_song in enumerate(sheet_songs):
        for s in all_songs:
            if matches(sheet_song, s):
                set_item = SetItem(s, order, sheet_song[1])
                matched_songs.append(set_item)
                break
    return matched_songs

def matches(sheet_song: tuple, song: Song):
    if sheet_song[0] in song.name:
        return True
    if sheet_song[0] == song.subtitle:
        return True
    return False

def validate_parameters(ip_address: str, spreadsheet: str, sheet: str, set_name: str):
    if not ip_address:
        raise ValueError("URL must not be empty.")
    if not spreadsheet:
        raise ValueError("Spreadsheet path must not be empty.")
    if not sheet:
        raise ValueError("Sheet name must not be empty.")
    if not set_name:
        raise ValueError("Set name must not be empty.")

def run(ip_address: str, spreadsheet: str, sheet: str, set_name: str):
    validate_parameters(ip_address, spreadsheet, sheet, set_name)

    song_access_object = SongbookProManagerService(ip_address)

    all_songs = song_access_object._find_all_songs()
    sheet_songs = load_songs_from_sheets(spreadsheet, sheet)
    matched_songs = match_songs(all_songs, sheet_songs)
    song_access_object.create_set(set_name, matched_songs)
