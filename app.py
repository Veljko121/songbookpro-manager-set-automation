import requests
import sys
import openpyxl as xl

base_url = "http://192.168.0.3:8080"

keys = {
    "A"  :  0,
    "B"  :  1,
    "H"  :  2,
    "C"  :  3,
    "Db" :  4,
    "D"  :  5,
    "Eb" :  6,
    "E"  :  7,
    "F"  :  8,
    "F#" :  9,
    "G"  : 10,
    "Ab" : 11,
    "F#m": 12,
    "Gm" : 13,
    "G#m": 14,
    "Am" : 15,
    "Bm" : 16,
    "Hm" : 17,
    "Cm" : 18,
    "C#m": 19,
    "Dm" : 20,
    "D#m": 21,
    "Em" : 22,
    "Fm" : 23,
}

class Song:
    def __init__(self, id: int, name: str, key: int, subtitle: str, key_shift: int):
        self.id = id
        self.name = name
        self.key = key
        self.subtitle = subtitle
        self.key_shift = key_shift

    def active_key(self):
        key = self.key + self.key_shift
        if self.key <= 11:
            key %= 12
        else:
            if key >= 24:
                key -= 12
        return key

class SetItem:
    def __init__(self, song: Song, order: int, set_key: int):
        self.song = song
        self.order = order
        self.set_key = set_key

    def key_offset(self):
        distance = 12 - (self.song.key - self.set_key) if self.song.key > self.set_key else self.set_key - self.song.key
        return distance

def load_songs_from_sheets(sheetsFilename: str, sheetName: str):
    try:
        try:
            doc = xl.load_workbook(sheetsFilename)
            sheet = doc[sheetName]
            print("Loading songs from '" + sheet.title + "'...")
            songs = [
                (str(sheet.cell(row, 2).value).replace("‘", "'").replace("’", "'"), keys[str(sheet.cell(row, 3).value)])
                for row in range(1, sheet.max_row + 1)
            ] # extremely complicated conversion of songs
            songs = [song for song in songs if song[0] != "None"] # removing None values
            return songs
        except FileNotFoundError:
            print("Spreadsheet '" + sheetsFilename + "' doesn't seem to exist. Try again.")
            exit(1)
    except KeyError:
        print("Worksheet '" + sheetName + "' doesn't seem to exist. Try again.")
        exit(1)

def fetch_songs(session: requests.Session, base_url: str):
    response = session.get(base_url + "/api/editor/songs")
    songs_json = response.json()
    songs = [Song(song['Id'], song['name'], song['key'], song['subTitle'], song['KeyShift']) for song in songs_json]
    return songs

def get_cookie(session: requests.Session, base_url: str):
    session.get(base_url + "/api/editor/songs")
    cookies = session.cookies.get_dict()
    cookie = cookies['jagses']
    return cookie

def match_songs(all_songs, sheet_songs):
    matched_songs = []
    for order, sheet_song in enumerate(sheet_songs):
        for s in all_songs:
            if sheet_song[0] == s.name:
                set_item = SetItem(s, order, sheet_song[1])
                matched_songs.append(set_item)
                break
            elif sheet_song[0] == s.subtitle:
                set_item = SetItem(s, order, sheet_song[1])
                matched_songs.append(set_item)
                break
    return matched_songs

def add_song(session: requests.Session, base_url: str, headers, set_id: int, set_item: SetItem):
    response = session.post(base_url + f"/api/arrange/setItem", headers=headers, json={"order": set_item.order, "setId": set_id, "songId": set_item.song.id, "type": 1})
    set_item_id = response.json()["setItem"]["Id"]
    response = session.put(base_url + "/api/arrange/setItem", headers=headers, json=[{"id": set_item_id, "data": {"keyOfset": set_item.key_offset()}}])

def create_set(session: requests.Session, base_url: str, headers, set_name: str, matched_songs):
    response = session.post(base_url + "/api/arrange/sets", headers=headers)
    set_id = response.json()["Id"]
    session.put(base_url + f"/api/arrange/sets/{set_id}", json={"name": set_name})
    for matched_song in matched_songs:
        add_song(session, base_url, headers, set_id, matched_song)

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Invalid number of arguments!")
        exit(1)
    session = requests.Session()
    cookie = get_cookie(session, base_url)
    headers = {
        "Cookie": f"jagses={cookie}"
    }
    sheets_filename = sys.argv[1]
    sheet_name = sys.argv[2]
    set_name = sys.argv[3]
    all_songs = fetch_songs(session, base_url)
    sheet_songs = load_songs_from_sheets(sheets_filename, sheet_name)
    matched_songs = match_songs(all_songs, sheet_songs)
    create_set(session, base_url, headers, set_name, matched_songs)