import requests
import openpyxl as xl

keys = {
    "A"  :  0,
    "B"  :  1,
    "H"  :  2,
    "C"  :  3,
    "Db" :  4, "C#" :  4,
    "D"  :  5,
    "Eb" :  6, "D#" :  6,
    "E"  :  7,
    "F"  :  8,
    "F#" :  9, "Gb" :  9,
    "G"  : 10,
    "Ab" : 11, "G#" : 11, 
    "F#m": 12, "Gbm": 12,
    "Gm" : 13,
    "G#m": 14, "Abm": 14,
    "Am" : 15,
    "Bm" : 16,
    "Hm" : 17,
    "Cm" : 18,
    "C#m": 19, "Dbm": 19,
    "Dm" : 20,
    "D#m": 21, "Ebm": 21,
    "Em" : 22,
    "Fm" : 23,

    "None" : -1,
}

class Song:
    def __init__(self, id: int, name: str, key: int, subtitle: str, key_shift: int):
        self.id = id
        self.name = name
        self.key = key
        self.subtitle = subtitle
        self.key_shift = key_shift

    def active_key(self):
        key = self.key + self.key_shift
        if self.key <= 11:
            key %= 12
        else:
            if key >= 24:
                key -= 12
        return key

class SetItem:
    def __init__(self, song: Song, order: int, set_key: int):
        self.song = song
        self.order = order
        self.set_key = set_key

    def key_offset(self):
        distance = 12 - (self.song.key - self.set_key) if self.song.key > self.set_key else self.set_key - self.song.key
        return distance
    
def get_sheet_names(spreadsheet: str):
    """Fetch all sheet names from the given spreadsheet."""
    return xl.load_workbook(spreadsheet, read_only=True).sheetnames


def load_songs_from_sheets(spreadsheet: str, sheet_name: str):
    """
    Load songs from the given spreadsheet and sheet name.
    ## Exceptions
    - FileNotFoundError
    - KeyError
    - ValueError
    """
    doc = xl.load_workbook(spreadsheet)
    sheet = doc[sheet_name]
    
    songs = []
    row = 1
    while True:
        name_cell = sheet.cell(row, 2).value  # Column B
        if name_cell is None:
            break  # Stop if the B column is empty
        
        # Normalize song name and load key
        name = str(name_cell).replace("‘", "'").replace("’", "'")
        key_cell = sheet.cell(row, 3).value  # Column C
        if key_cell is None:
            raise ValueError(f"Missing key for song '{name}' in row {row}")
        
        key_str = str(key_cell).strip()
        if key_str not in keys:
            raise KeyError(f"Unknown key '{key_str}' for song '{name}' in row {row}")
        
        key_value = keys[key_str]
        songs.append((name, key_value))
        row += 1

    return songs

def fetch_songs(session: requests.Session, base_url: str):
    """
    Fetch all songs from SongbookPro.
    """
    response = session.get(base_url + "/api/editor/songs")
    songs_json = response.json()
    songs = [Song(song['Id'], song['name'], song['key'], song['subTitle'], song['KeyShift']) for song in songs_json]
    return songs

def get_cookie(session: requests.Session, base_url: str):
    print('Waiting for approval from application...')
    session.get(base_url + "/api/editor/songs")
    cookies = session.cookies.get_dict()
    cookie = cookies['jagses']
    return cookie

def match_songs(all_songs, sheet_songs):
    matched_songs = []
    for order, sheet_song in enumerate(sheet_songs):
        for s in all_songs:
            if matches(sheet_song, s):
                set_item = SetItem(s, order, sheet_song[1])
                matched_songs.append(set_item)
                break
    return matched_songs

def matches(sheet_song: tuple, song: Song):
    if sheet_song[0] in song.name:
        return True
    if sheet_song[0] == song.subtitle:
        return True
    return False

def add_song(session: requests.Session, base_url: str, headers, set_id: int, set_item: SetItem):
    response = session.post(base_url + f"/api/arrange/setItem", headers=headers, json={"order": set_item.order, "setId": set_id, "songId": set_item.song.id, "type": 1})
    set_item_id = response.json()["setItem"]["Id"]
    if set_item.set_key > 0:
        response = session.put(base_url + "/api/arrange/setItem", headers=headers, json=[{"id": set_item_id, "data": {"keyOfset": set_item.key_offset()}}])

def create_set(session: requests.Session, base_url: str, headers, set_name: str, matched_songs):
    response = session.post(base_url + "/api/arrange/sets", headers=headers)
    set_id = response.json()["Id"]
    session.put(base_url + f"/api/arrange/sets/{set_id}", json={"name": set_name})
    for matched_song in matched_songs:
        add_song(session, base_url, headers, set_id, matched_song)
    print(f"Set '{set_name}' created successfully!")

def run(ipAddress: str, spreadsheet: str, sheet: str, set_name: str):
    # Check if any argument is empty and raise specific errors
    if not ipAddress:
        raise ValueError("URL must not be empty.")
    if not spreadsheet:
        raise ValueError("Spreadsheet path must not be empty.")
    if not sheet:
        raise ValueError("Sheet name must not be empty.")
    if not set_name:
        raise ValueError("Set name must not be empty.")
    
    url = f"http://{ipAddress}:8080"

    session = requests.Session()
    cookie = get_cookie(session, url)
    if not cookie:
        print("Failed to retrieve session cookie.")
        exit(1)

    headers = {"Cookie": f"jagses={cookie}"}

    all_songs = fetch_songs(session, url)
    sheet_songs = load_songs_from_sheets(spreadsheet, sheet)
    matched_songs = match_songs(all_songs, sheet_songs)
    create_set(session, url, headers, set_name, matched_songs)
